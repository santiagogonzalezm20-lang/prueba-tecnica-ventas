from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

RUTA = "output/resumen_ventas.xlsx"

wb = load_workbook(RUTA)

header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
body_font = Font(name="Calibri", size=11)

for hoja in wb.sheetnames:
    ws = wb[hoja]

    # Encabezados
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    # Cuerpo
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=row, column=col)
            c.font = body_font
            header = ws.cell(row=1, column=col).value
            if header and "Total" in str(header):
                c.number_format = '"$"#,##0'

    # Ancho de columnas
    for col in range(1, ws.max_column + 1):
        letra = get_column_letter(col)
        ws.column_dimensions[letra].width = 18

    ws.freeze_panes = "A2"

wb.save(RUTA)
print("Formato aplicado")
